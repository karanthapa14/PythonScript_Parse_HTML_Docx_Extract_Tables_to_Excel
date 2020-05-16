#!/usr/bin/env python

# Make necessary imports
import tkinter
from pathlib import Path
from tkinter import filedialog

from bs4 import BeautifulSoup as soup
from docx import Document
from docx.document import Document as _Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


# Find and return T-Codes from Flow Blocks in the HTML File
# @params page_soup bs4.BeautifulSoup object
def get_tcode_html(page_soup):
    tcodes = []
    containers = page_soup.findAll("text", {"class": "taskName"})
    for container in containers:
        str = "".join(e for e in container.text if e.isalnum()).lower()
        tcodes.append(str)
    return tcodes


# Get reference to each paragraph and table child within parent, in document order.
# @params parent Document object
def iter_block_docx(parent):
    if isinstance(parent, _Document):
        parent_elm = parent.element.body
    else:
        raise ValueError("Something went right")

    for child in parent_elm.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, parent)
        elif isinstance(child, CT_Tbl):
            yield Table(child, parent)


# Extract valid table from Docx file and append it to new Workbook
# @params dir_save Directory path, workbookname Filename, doc Document obj, tcodes List of Tcodes in HTML
def crate_workbook(dir_save, workbookname, doc, tcodes):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False

    heading = False
    procedure = False
    tCode = None

    medium_border = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='medium'),
                           bottom=Side(style='medium'))

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    medium_bottom_border = Border(bottom=Side(style='medium'),
                                  right=Side(style='thin'))

    headers = ["T-Code", "Test Step #", "Test Step Name", "Instruction", "Expected Result", "Pass / Fail / Comment"]
    for cell_col_no, data in enumerate(headers):
        cell = ws.cell(row=1, column=cell_col_no + 1, value=data)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = medium_border

    cell_row_no = 2
    for block in iter_block_docx(doc):

        if isinstance(block, Paragraph):

            if block.style.name.startswith("Heading"):
                if ''.join(e for e in block.text if e.isalnum()).lower() in tcodes:
                    heading = True
                    tCode = block.text
                else:
                    heading = False
                    tCode = None

            if heading and block.text == "Procedure":
                procedure = True
            else:
                procedure = False

        elif heading and procedure and isinstance(block, Table):

            cell = ws.cell(row=cell_row_no, column=1, value=tCode)
            cell.alignment = Alignment(horizontal="center", vertical="top")
            cell.font = Font(bold=True)
            cell.border = medium_border

            start_row = cell_row_no
            table_total_rows = len(block.rows[1:])
            for table_row_no, table_row in enumerate(block.rows[1:]):

                if table_row.cells[0].text == "" or table_row.cells[0].text.isnumeric():

                    for table_cell_no, table_cell in enumerate(table_row.cells):
                        para_list = []
                        for paragraph in table_cell.paragraphs:
                            para_list.append(paragraph.text)
                        actual_para = "\n".join(para_list)
                        cell = ws.cell(row=cell_row_no, column=table_cell_no + 2, value=actual_para)
                        cell.border = thin_border

                        if table_row_no == table_total_rows - 1:
                            cell.border = medium_bottom_border
                        if table_cell_no == 0:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.font = Font(bold=True)
                        elif table_cell_no == 1:
                            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

                else:
                    cell = ws.cell(row=cell_row_no, column=2, value=table_row.cells[0].text)
                    ws.merge_cells(start_row=cell_row_no, start_column=2, end_row=cell_row_no, end_column=6)

                cell_row_no = cell_row_no + 1

            ws.merge_cells(start_row=start_row, start_column=1, end_row=cell_row_no - 1, end_column=1)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 120
    ws.column_dimensions["E"].width = 120
    ws.column_dimensions["F"].width = 20

    wb.save(dir_save + workbookname + '.xlsx')
    print("Workbook Created: " + "'" + dir_save + workbookname + '.xlsx' + "'")


# Entry Point
# hide main window
root = tkinter.Tk()
root.withdraw()

input_html_dir = filedialog.askopenfilename(initialdir='inputFiles', title="Select HTML file",
                                            filetypes=(("HTML files", "*.htm?"), ("All files", "*.*")))

ext1 = Path(input_html_dir).suffix
if not (ext1 == ".html" or ext1 == ".htm"):
    raise Exception("Select HTML File, try agian!")

input_docx_dir = filedialog.askopenfilename(initialdir='inputFiles', title="Select Docx file",
                                            filetypes=(("Doc files", "*.doc?"), ("All files", "*.*")))

ext2 = Path(input_docx_dir).suffix
if not (ext2 == ".docx" or ext1 == ".doc"):
    raise Exception("Select Document File, try agian!")

fileName1 = Path(input_html_dir).stem
fileName2 = Path(input_docx_dir).stem

with open(input_html_dir, 'r') as html_file:
    html_soup = soup(html_file, "html.parser")

with open(input_docx_dir, 'r'):
    doc_obj = Document(input_docx_dir)

list_tcodes = get_tcode_html(html_soup)

excel_dir = filedialog.askdirectory(initialdir='outputFiles', title="Select Folder to save Excel sheet")
excel_dir = excel_dir + "/"

crate_workbook(excel_dir, fileName2, doc_obj, list_tcodes)